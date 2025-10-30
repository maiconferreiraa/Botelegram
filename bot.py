# -*- coding: utf-8 -*-
import os
import re
import io
import decimal
from decimal import Decimal
from datetime import datetime, timedelta
import asyncio

# Imports do Bot
from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import Application, MessageHandler, filters, ContextTypes, CommandHandler

# Imports dos GrÃ¡ficos/RelatÃ³rios
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

# --- Import do Flask e Thread (NecessÃ¡rio para Render/Replit) ---
from flask import Flask
from threading import Thread
# ----------------------------------------

from db import db # importa a instÃ¢ncia do db.py

# =======================
# CONFIGURAÃ‡ÃƒO ADMIN
# =======================
ADMIN_USER_ID = 853716041 # ID @maiconjbf

# ===================================================================
# --- MAPEAMENTO DE CATEGORIAS (Como definido anteriormente) ---
# ===================================================================
MAPEAMENTO_CATEGORIAS = {
    # --- GASTOS ---
    "AlimentaÃ§Ã£o": ["supermercado", "mercado", "lanche", "churrasco", "restaurante", "ifood", "rappi", "padaria", "aÃ§ougue", "hortifruti", "pizza", "comida", "jantar", "almoÃ§o", "cafÃ©", "bebida"],
    "Transporte": ["gasolina", "uber", "99", "estacionamento", "ipva", "seguro", "carro", "manutenÃ§Ã£o", "onibus", "metrÃ´", "passagem", "combustÃ­vel", "pedagio", "taxi", "aplicativo", "app"],
    "Moradia": ["aluguel", "condomÃ­nio", "iptu", "luz", "Ã¡gua", "internet", "gÃ¡s", "diarista", "faxina", "energia", "net", "claro", "vivo", "oi", "tim", "conserto", "reparo", "internet celular", "celular internet"],
    "ConstruÃ§Ã£o/Reforma": ["construÃ§Ã£o", "reforma", "material", "pedreiro", "tinta", "cimento", "leroy", "telhanorte", "ferramenta", "obra", "ferragens"],
    "Casa/DecoraÃ§Ã£o": ["casa", "decoraÃ§Ã£o", "mÃ³vel", "utensÃ­lio", "cama", "mesa", "banho", "eletrodomÃ©stico", "manutenÃ§Ã£o", "casa", "jardinagem", "ikea", "tokstok"],
    "SaÃºde": ["farmÃ¡cia", "remÃ©dio", "mÃ©dico", "consulta", "plano", "saude", "exame", "dentista", "hospital", "terapia", "psicologo"],
    "Lazer/Entretenimento": ["lazer", "cinema", "show", "bar", "festa", "viagem", "hotel", "streaming", "netflix", "spotify", "hobby", "jogo", "steam", "passeio", "presente", "ingresso", "assinatura", "disney", "hbo"],
    "EducaÃ§Ã£o": ["escola", "faculdade", "curso", "livro", "material", "escolar", "udemy", "mensalidade", "papelaria"],
    "VestuÃ¡rio/Cuidados": ["roupa", "sapato", "tÃªnis", "acessÃ³rio", "vestido", "calÃ§a", "beleza", "cabelereiro", "cosmÃ©tico", "perfume", "barbeiro"],
    "DÃ­vidas/Contas": ["fatura", "emprÃ©stimo", "juros", "boleto", "imposto", "taxa", "ir", "multa", "cartorio"],
    "Pets": ["pet", "raÃ§Ã£o", "veterinÃ¡rio", "petshop", "cachorro", "gato"],
    # --- ENTRADAS ---
    "SalÃ¡rio": ["salÃ¡rio", "salario", "pagamento", "holerite"],
    "Vendas": ["venda", "cliente", "recebimento", "comissao", "serviÃ§o", "cliente pagou"],
    "Investimentos": ["investimento", "aÃ§Ã£o", "aÃ§Ãµes", "b3", "fundo", "tesouro", "cdb", "cripto", "resgate", "dividendo", "jcp"],
    "Outras Entradas": ["entrada", "ganhei", "recebi", "pix", "reembolso", "presente"]
}

# =======================
# --- FUNÃ‡ÃƒO HELPER (Como definido anteriormente) ---
# =======================
def encontrar_categoria_por_palavra(palavras: list):
    for palavra in palavras:
        for categoria_pai, keywords in MAPEAMENTO_CATEGORIAS.items():
            if palavra in keywords: return categoria_pai
    return None

# =======================
# FunÃ§Ã£o para formatar valores BR
# =======================
def formatar_valor(valor):
    try: valor_decimal = Decimal(valor)
    except (decimal.InvalidOperation, TypeError, ValueError): valor_decimal = Decimal("0.00")
    return f"{valor_decimal:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

# =======================
# InterpretaÃ§Ã£o de mensagens (MODIFICADA)
# =======================
def interpretar_mensagem(texto: str):
    texto = texto.lower().strip()
    match = re.search(r"(\d[\d.,]*)", texto)
    if match:
        valor_txt = match.group(1).strip(); valor_num = Decimal("0.00")
        if not valor_txt: return {"acao": "desconhecido"}
        try: valor_num = Decimal(valor_txt.replace(".", "").replace(",", "."))
        except decimal.InvalidOperation: return {"acao": "desconhecido"}
        if valor_num <= 0: return {"acao": "desconhecido"}

        palavras = texto.split()
        palavras_texto = [p for p in palavras if valor_txt not in p]

        # --- LÃ³gica de CartÃ£o ---
        cartao = None; metodo = "dinheiro"; cartoes_lista = ["nubank", "santander", "inter", "caixa"]; stop_words_cartao = cartoes_lista + ["cartÃ£o", "cartao"]
        for c in cartoes_lista:
            if c in palavras_texto: cartao = c.capitalize(); metodo = "cartao"; break
        if cartao is None:
            idx = -1
            if "cartÃ£o" in palavras_texto: idx = palavras_texto.index("cartÃ£o")
            elif "cartao" in palavras_texto: idx = palavras_texto.index("cartao")
            if idx != -1:
                metodo = "cartao"; nome_cartao_palavras = []
                temp_stop_words = stop_words_cartao + list(MAPEAMENTO_CATEGORIAS.keys())
                for i in range(idx + 1, len(palavras_texto)):
                    if palavras_texto[i] not in temp_stop_words: nome_cartao_palavras.append(palavras_texto[i])
                    else: break
                if nome_cartao_palavras: cartao = " ".join(nome_cartao_palavras).capitalize()
                else: cartao = "CartÃ£o"
        # --- Fim CartÃ£o ---

        # --- Determina Tipo e Categoria ---
        entradas_keywords = [kw for cat, kws in MAPEAMENTO_CATEGORIAS.items() if cat in ["SalÃ¡rio", "Vendas", "Outras Entradas", "Investimentos"] for kw in kws]
        is_entrada = any(p in entradas_keywords for p in palavras_texto)

        if is_entrada:
            categoria_mapeada = encontrar_categoria_por_palavra(palavras_texto)
            categoria_final = categoria_mapeada if categoria_mapeada else "Entrada"
            return {"acao": "add", "tipo": "entrada", "valor_num": valor_num, "valor_txt": valor_txt, "categoria": categoria_final.capitalize(), "metodo": metodo, "cartao": cartao}
        else:
            categoria_mapeada = encontrar_categoria_por_palavra(palavras_texto)
            categoria_final = None
            if categoria_mapeada:
                categoria_final = categoria_mapeada
            else:
                stop_words_fallback = stop_words_cartao
                if cartao: stop_words_fallback.extend(cartao.lower().split())
                categoria_primeira_palavra = next((p for p in palavras_texto if p.isalpha() and p not in stop_words_fallback), "Outros")
                categoria_final = categoria_primeira_palavra.capitalize()
            return {"acao": "add", "tipo": "gasto", "valor_num": valor_num, "valor_txt": valor_txt, "categoria": categoria_final.capitalize(), "metodo": metodo, "cartao": cartao}
    return {"acao": "desconhecido"}


# ==========================================================
# --- MODIFICAÃ‡ÃƒO 1: Emojis e BotÃµes do Teclado Flutuante ---
# ==========================================================
def teclado_flutuante(user_id):
    entradas = db.get_soma(user_id, "entrada"); gastos = db.get_soma(user_id, "gasto"); saldo = entradas - gastos
    status = "ğŸŸ¢ğŸ˜€ FinanÃ§as SaudÃ¡veis"
    if saldo < 0: status = "ğŸ”´ğŸ˜Ÿ Saldo Negativo"
    elif entradas > 0 and (gastos / entradas) > Decimal("0.7"): status = "ğŸŸ ğŸ¤” Gastos altos!"
    
    # Emojis Ãºnicos para cada botÃ£o
    teclado = [
        [status],
        ["âš–ï¸ Saldo Geral", "ğŸ’³ Gastos por CartÃ£o"],
        ["ğŸ“¥ Ver Entradas", "ğŸ“¤ Ver SaÃ­das"],
        ["ğŸ—“ï¸ Filtrar por PerÃ­odo", "ğŸ·ï¸ Filtrar por Categoria"],
        ["ğŸ• GrÃ¡fico Pizza", "ğŸ“Š GrÃ¡fico Barras"],
        ["ğŸ“„ Gerar PDF", "ğŸ“ˆ Gerar XLSX", "ğŸ—‘ï¸ Resetar Valores"],
        ["ğŸ¤– Quero um robÃ´"] # <-- NOVO BOTÃƒO
    ]
    if user_id == ADMIN_USER_ID: 
        teclado.append(["ğŸ§‘â€ğŸ’¼ Ver UsuÃ¡rios"]) # Emoji de Admin
    return ReplyKeyboardMarkup(teclado, resize_keyboard=True, one_time_keyboard=False)

def teclado_admin_usuario_selecionado():
    teclado = [["ğŸ’° Entradas", "ğŸ’¸ SaÃ­das"], ["ğŸ§¾ Saldo Geral"], ["ğŸ“‘ Gerar PDF", "ğŸ“Š Gerar XLSX"], ["â¬…ï¸ Voltar"]]
    return ReplyKeyboardMarkup(teclado, resize_keyboard=True, one_time_keyboard=False)

def teclado_filtros_periodo():
    teclado = [["Hoje", "Esta Semana", "Este MÃªs"], ["MÃªs Passado", "Este Ano"], ["Cancelar"]]
    return ReplyKeyboardMarkup(teclado, resize_keyboard=True, one_time_keyboard=True)

# =======================
# FunÃ§Ãµes de GrÃ¡ficos, PDF, XLSX, etc. (Sem alteraÃ§Ãµes)
# =======================
def grafico_gastos_pizza(user_id=None, inicio=None, fim=None):
    rows = db.gastos_por_categoria(user_id=user_id, inicio=inicio, fim=fim)
    if not rows: return None
    labels = [r[0] for r in rows]; valores = [float(r[1]) for r in rows]; fig, ax = plt.subplots()
    ax.pie(valores, labels=labels, autopct="%1.1f%%", startangle=90); ax.set_title("Gastos por Categoria")
    buf = io.BytesIO(); plt.savefig(buf, format="png", bbox_inches="tight"); buf.seek(0); plt.close(fig); return buf

def grafico_mensal_barras(user_id=None, meses=6):
    labels, entradas_vals, gastos_vals = db.series_mensais(user_id=user_id, meses=meses)
    if not labels: return None
    x = list(range(len(labels))); fig, ax = plt.subplots(); width = 0.4
    ax.bar([i - width/2 for i in x], entradas_vals, width=width, label="Entradas", align="center")
    ax.bar([i + width/2 for i in x], gastos_vals, width=width, label="Gastos", align="center")
    ax.set_xticks(x); ax.set_xticklabels(labels, rotation=45); ax.set_ylabel("R$")
    ax.set_title("Entradas x Gastos por MÃªs"); ax.legend(); fig.tight_layout()
    buf = io.BytesIO(); plt.savefig(buf, format="png", bbox_inches="tight"); buf.seek(0); plt.close(fig); return buf

def gerar_pdf(user_id=None, filename="relatorio.pdf", inicio=None, fim=None):
    doc = SimpleDocTemplate(filename); styles = getSampleStyleSheet(); story = []
    entradas = db.get_soma(user_id, "entrada", inicio=inicio, fim=fim); gastos = db.get_soma(user_id, "gasto", inicio=inicio, fim=fim); saldo = entradas - gastos
    story.append(Paragraph("ğŸ“‘ RelatÃ³rio Financeiro", styles["Title"])); story.append(Spacer(1, 20))
    story.append(Paragraph(f"Entradas: R$ {formatar_valor(entradas)}", styles["Normal"])); story.append(Paragraph(f"Gastos: R$ {formatar_valor(gastos)}", styles["Normal"]))
    story.append(Paragraph(f"Saldo: R$ {formatar_valor(saldo)}", styles["Normal"])); story.append(Spacer(1, 20))
    story.append(Paragraph("ğŸ’° Entradas:", styles["Heading2"]))
    trans_e = db.get_todas(user_id=user_id, tipo="entrada", inicio=inicio, fim=fim)
    for t in trans_e: story.append(Paragraph(f"â¡ï¸ R$ {formatar_valor(t[2])} ({t[3]}) - {t[5] or 'Dinheiro'} - {t[6]}", styles["Normal"]))
    story.append(Spacer(1, 20)); story.append(Paragraph("ğŸ’¸ SaÃ­das:", styles["Heading2"]))
    trans_s = db.get_todas(user_id=user_id, tipo="gasto", inicio=inicio, fim=fim)
    for t in trans_s: story.append(Paragraph(f"â¬…ï¸ R$ {formatar_valor(t[2])} ({t[3]}) - {t[5] or 'Dinheiro'} - {t[6]}", styles["Normal"]))
    doc.build(story); return filename

def gerar_xlsx(user_id=None, filename="relatorio.xlsx", inicio=None, fim=None):
    wb = Workbook(); ws = wb.active; ws.title = "RelatÃ³rio"; ws.append(["Tipo", "Valor", "Categoria", "MÃ©todo", "CartÃ£o", "Data"])
    transacoes = db.get_todas(user_id=user_id, inicio=inicio, fim=fim)
    for t in transacoes:
        try: valor_num = Decimal(t[2])
        except (decimal.InvalidOperation, TypeError, ValueError): valor_num = Decimal("0.00")
        ws.append([t[1], valor_num, t[3], t[4], t[5] or "Dinheiro", t[6]])
    entradas = db.get_soma(user_id, "entrada", inicio=inicio, fim=fim); gastos = db.get_soma(user_id, "gasto", inicio=inicio, fim=fim); saldo = entradas - gastos
    ws.append([]); ws.append(["Entradas", entradas]); ws.append(["Gastos", gastos]); ws.append(["Saldo", saldo])
    num_format = 'R$ #,##0.00'; max_row = ws.max_row
    for cell in ws['B']: cell.number_format = num_format
    ws[f'B{max_row-2}'].number_format = num_format; ws[f'B{max_row-1}'].number_format = num_format; ws[f'B{max_row}'].number_format = num_format
    wb.save(filename); return filename

def gastos_por_cartao(user_id):
    rows = db.get_gastos_por_cartao(user_id=user_id)
    if not rows: return "ğŸ’³ Gastos por CartÃ£o:\nNenhum gasto registrado."
    texto = "ğŸ’³ Gastos por CartÃ£o:\n";
    for r in rows: texto += f"â–ªï¸ {r[0]}: R$ {formatar_valor(r[1])}\n"
    return texto

def verificar_alerta(user_id):
    entradas = db.get_soma(user_id, "entrada"); gastos = db.get_soma(user_id, "gasto"); saldo = entradas - gastos
    status = None
    if saldo < 0: status = "ğŸ”´ğŸ˜Ÿ Saldo Negativo"
    elif entradas > 0 and (gastos / entradas) > Decimal("0.7"): status = "ğŸŸ ğŸ¤” Gastos altos!"
    if status: return (f"{status}\nğŸ’° Entradas: R$ {formatar_valor(entradas)}\nğŸ’¸ Gastos: R$ {formatar_valor(gastos)}\nğŸ“Œ Saldo: R$ {formatar_valor(saldo)}")
    return None

async def enviar_extrato_filtrado(update: Update, context: ContextTypes.DEFAULT_TYPE, inicio: datetime, fim: datetime, titulo_periodo: str):
    user_id = update.message.from_user.id
    entradas = db.get_todas(user_id, tipo="entrada", inicio=inicio, fim=fim)
    saidas = db.get_todas(user_id, tipo="gasto", inicio=inicio, fim=fim)
    entradas_filtradas = [t for t in entradas if Decimal(t[2]) > 0]; saidas_filtradas = [t for t in saidas if Decimal(t[2]) > 0]
    total_entradas = db.get_soma(user_id, "entrada", inicio=inicio, fim=fim); total_gastos = db.get_soma(user_id, "gasto", inicio=inicio, fim=fim); saldo_periodo = total_entradas - total_gastos
    texto = f"ğŸ§¾ Extrato Filtrado: *{titulo_periodo}*\n\n"
    if not entradas_filtradas and not saidas_filtradas: texto += "Nenhuma transaÃ§Ã£o neste perÃ­odo."
    else:
        if entradas_filtradas:
            texto += "--- *Entradas* ---\n";
            for t in entradas_filtradas: texto += f"â¡ï¸ R$ {formatar_valor(t[2])} ({t[3]}) - {t[6]}\n"
            texto += "\n"
        if saidas_filtradas:
            texto += "--- *SaÃ­das* ---\n"
            for t in saidas_filtradas: texto += f"â¬…ï¸ R$ {formatar_valor(t[2])} ({t[3]}) - {t[5] or 'Dinheiro'} - {t[6]}\n"
            texto += "\n"
        texto += "--- *Resumo do PerÃ­odo* ---\n"; texto += f"ğŸ’° Total Entradas: R$ {formatar_valor(total_entradas)}\n"; texto += f"ğŸ’¸ Total Gastos: R$ {formatar_valor(total_gastos)}\n"; texto += f"ğŸ“Œ Saldo PerÃ­odo: R$ {formatar_valor(saldo_periodo)}\n"
    await update.message.reply_text(texto, parse_mode='Markdown', reply_markup=teclado_flutuante(user_id))

# =======================
# FunÃ§Ã£o de Filtro por Categoria (JÃ¡ incluÃ­da)
# =======================
async def enviar_extrato_por_categoria(update: Update, context: ContextTypes.DEFAULT_TYPE, categoria_desejada: str):
    user_id = update.message.from_user.id
    categoria_lower = categoria_desejada.lower().strip()
    entradas_todas = db.get_todas(user_id, tipo="entrada")
    saidas_todas = db.get_todas(user_id, tipo="gasto")
    entradas_filtradas = [t for t in entradas_todas if t[3].lower() == categoria_lower and Decimal(t[2]) > 0]
    saidas_filtradas = [t for t in saidas_todas if t[3].lower() == categoria_lower and Decimal(t[2]) > 0]
    total_entradas = sum(Decimal(t[2]) for t in entradas_filtradas)
    total_gastos = sum(Decimal(t[2]) for t in saidas_filtradas)
    saldo_categoria = total_entradas - total_gastos
    texto = f"ğŸ§¾ Extrato Filtrado: *Categoria: {categoria_desejada.capitalize()}*\n\n"
    if not entradas_filtradas and not saidas_filtradas:
        texto += "Nenhuma transaÃ§Ã£o encontrada para esta categoria."
        await update.message.reply_text(texto, parse_mode='Markdown', reply_markup=teclado_flutuante(user_id))
        return
    if entradas_filtradas:
        texto += "--- *Entradas* ---\n"
        for t in entradas_filtradas: texto += f"â¡ï¸ R$ {formatar_valor(t[2])} ({t[3]}) - {t[6]}\n"
        texto += "\n"
    if saidas_filtradas:
        texto += "--- *SaÃ­das* ---\n"
        for t in saidas_filtradas: texto += f"â¬…ï¸ R$ {formatar_valor(t[2])} ({t[3]}) - {t[5] or 'Dinheiro'} - {t[6]}\n"
        texto += "\n"
    texto += f"--- *Resumo da Categoria: {categoria_desejada.capitalize()}* ---\n"
    texto += f"ğŸ’° Total Entradas: R$ {formatar_valor(total_entradas)}\n"; texto += f"ğŸ’¸ Total Gastos: R$ {formatar_valor(total_gastos)}\n"; texto += f"ğŸ“Œ Saldo Categoria: R$ {formatar_valor(saldo_categoria)}\n"
    await update.message.reply_text(texto, parse_mode='Markdown', reply_markup=teclado_flutuante(user_id))

# =======================
# Handlers
# =======================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id; user_name = update.message.from_user.first_name
    await update.message.reply_text(f"OlÃ¡, {user_name}! Bem-vindo(a).\n"
                                     "Digite valor + descriÃ§Ã£o (ex: '150 mercado').\n"
                                     "Use o teclado para outras opÃ§Ãµes:",
                                     reply_markup=teclado_flutuante(user_id))

# ==========================================================
# --- MODIFICAÃ‡ÃƒO 2: FunÃ§Ã£o Responder (Atualizada) ---
# ==========================================================
async def responder(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id; user_name = update.message.from_user.first_name
    msg = update.message.text

    # --- Bloco 1: Capturar resposta do filtro por categoria ---
    if 'aguardando_filtro_categoria' in context.user_data:
        del context.user_data['aguardando_filtro_categoria'] 
        categoria_digitada = msg.strip()
        if categoria_digitada.lower() != "cancelar":
            await enviar_extrato_por_categoria(update, context, categoria_digitada)
        else:
             await update.message.reply_text("Filtro por categoria cancelado.", reply_markup=teclado_flutuante(user_id))
        return 
    # --- Fim Bloco 1 ---

    # --- Handlers Voltar/Cancelar padrÃ£o ---
    if msg == "â¬…ï¸ Voltar" and user_id == ADMIN_USER_ID:
        if "admin_selecionado" in context.user_data: del context.user_data["admin_selecionado"]
        await update.message.reply_text("Voltando...", reply_markup=teclado_flutuante(user_id)); return
    if msg == "Cancelar":
        if 'aguardando_filtro' in context.user_data: del context.user_data['aguardando_filtro']
        if 'aguardando_filtro_categoria' in context.user_data: del context.user_data['aguardando_filtro_categoria']
        await update.message.reply_text("AÃ§Ã£o cancelada.", reply_markup=teclado_flutuante(user_id)); return

    # --- Bloco Admin (Sem alteraÃ§Ãµes) ---
    if user_id == ADMIN_USER_ID and "admin_selecionado" in context.user_data:
        selecionado_id, selecionado_nome = context.user_data["admin_selecionado"]
        if 'aguardando_filtro' in context.user_data: del context.user_data['aguardando_filtro']
        # (O cÃ³digo do admin foi omitido aqui para focar nas mudanÃ§as do usuÃ¡rio, mas ele permanece)
        if msg == "ğŸ’° Entradas":
            transacoes = db.get_todas(user_id=selecionado_id, tipo="entrada"); filtradas = [t for t in transacoes if Decimal(t[2]) > 0]
            texto = f"ğŸ’° Entradas de {selecionado_nome}\n" + "\n".join([f"â¡ï¸ R$ {formatar_valor(t[2])} ({t[3]}) - {t[5] or 'Dinheiro'} - {t[6]}" for t in filtradas]);
            if not filtradas: texto = f"{selecionado_nome} nÃ£o tem entradas."; await update.message.reply_text(texto, reply_markup=teclado_admin_usuario_selecionado())
        elif msg == "ğŸ’¸ SaÃ­das":
            transacoes = db.get_todas(user_id=selecionado_id, tipo="gasto"); filtradas = [t for t in transacoes if Decimal(t[2]) > 0]
            texto = f"ğŸ’¸ SaÃ­das de {selecionado_nome}\n" + "\n".join([f"â¬…ï¸ R$ {formatar_valor(t[2])} ({t[3]}) - {t[5] or 'Dinheiro'} - {t[6]}" for t in filtradas]);
            if not filtradas: texto = f"{selecionado_nome} nÃ£o tem saÃ­das."; await update.message.reply_text(texto, reply_markup=teclado_admin_usuario_selecionado())
        elif msg == "ğŸ§¾ Saldo Geral":
            entradas = db.get_soma(selecionado_id, "entrada"); gastos = db.get_soma(selecionado_id, "gasto"); saldo = entradas - gastos
            await update.message.reply_text(f"Saldo de {selecionado_nome}\nğŸ’° Entradas: R$ {formatar_valor(entradas)}\nğŸ’¸ Gastos: R$ {formatar_valor(gastos)}\nğŸ“Œ Saldo: R$ {formatar_valor(saldo)}", reply_markup=teclado_admin_usuario_selecionado())
        elif msg == "ğŸ“‘ Gerar PDF": filename = gerar_pdf(selecionado_id, f"rel_{selecionado_id}.pdf"); await update.message.reply_document(open(filename, "rb"), caption=f"PDF de {selecionado_nome}", reply_markup=teclado_admin_usuario_selecionado()); os.remove(filename)
        elif msg == "ğŸ“Š Gerar XLSX": filename = gerar_xlsx(selecionado_id, f"rel_{selecionado_id}.xlsx"); await update.message.reply_document(open(filename, "rb"), caption=f"XLSX de {selecionado_nome}", reply_markup=teclado_admin_usuario_selecionado()); os.remove(filename)
        else: await update.message.reply_text("InvÃ¡lido.", reply_markup=teclado_admin_usuario_selecionado())
        return

    # --- Resposta Filtro PerÃ­odo (Atualizado com novo nome) ---
    if 'aguardando_filtro' in context.user_data:
        del context.user_data['aguardando_filtro']; hoje = datetime.now()
        if msg == "Hoje": inicio = fim = hoje
        elif msg == "Esta Semana": inicio = hoje - timedelta(days=hoje.weekday()); fim = inicio + timedelta(days=6)
        elif msg == "Este MÃªs": inicio = hoje.replace(day=1); fim = (inicio + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        elif msg == "MÃªs Passado": fim = hoje.replace(day=1) - timedelta(days=1); inicio = fim.replace(day=1)
        elif msg == "Este Ano": inicio = hoje.replace(day=1, month=1); fim = hoje.replace(day=31, month=12)
        else: await update.message.reply_text("Filtro cancelado.", reply_markup=teclado_flutuante(user_id)); return
        await enviar_extrato_filtrado(update, context, inicio, fim, msg); return

    # --- LÃ³gica UsuÃ¡rio Comum (Atualizada com novos emojis/nomes) ---
    if msg == "ğŸ—‘ï¸ Resetar Valores": 
        await update.message.reply_text("PerÃ­odo para resetar:", reply_markup=ReplyKeyboardMarkup([["Ãšltimo valor", "Hoje"], ["Ãšltima semana", "Este mÃªs"], ["Tudo"], ["Cancelar"]], resize_keyboard=True, one_time_keyboard=True)); return
    elif msg in ["Ãšltimo valor", "Hoje", "Ãšltima semana", "Este mÃªs", "Tudo"]: 
        mapa = {"Ãšltimo valor":"ultimo","Hoje":"dia","Ãšltima semana":"semana","Este mÃªs":"mes","Tudo":"tudo"}; db.limpar_transacoes(user_id, mapa[msg]); await update.message.reply_text(f"âœ… Removido ({msg})", reply_markup=teclado_flutuante(user_id)); return

    if msg == "ğŸ• GrÃ¡fico Pizza": 
        buf = grafico_gastos_pizza(user_id); await update.message.reply_photo(buf, caption="ğŸ’¸ Gastos por Categoria", reply_markup=teclado_flutuante(user_id)) if buf else await update.message.reply_text("Nenhum gasto.", reply_markup=teclado_flutuante(user_id)); return
    if msg == "ğŸ“Š GrÃ¡fico Barras": 
        buf = grafico_mensal_barras(user_id); await update.message.reply_photo(buf, caption="ğŸ“Š Entradas x Gastos", reply_markup=teclado_flutuante(user_id)) if buf else await update.message.reply_text("Nenhuma transaÃ§Ã£o.", reply_markup=teclado_flutuante(user_id)); return

    if msg == "ğŸ“¥ Ver Entradas": 
        transacoes = db.get_todas(user_id=user_id, tipo="entrada"); filtradas = [t for t in transacoes if Decimal(t[2]) > 0]; await update.message.reply_text("Nenhuma entrada.", reply_markup=teclado_flutuante(user_id)) if not filtradas else await update.message.reply_text("ğŸ’° Entradas:\n" + "\n".join([f"â¡ï¸ R$ {formatar_valor(t[2])} ({t[3]}) - {t[6]}" for t in filtradas]), reply_markup=teclado_flutuante(user_id)); return
    if msg == "ğŸ“¤ Ver SaÃ­das": 
        transacoes = db.get_todas(user_id=user_id, tipo="gasto"); filtradas = [t for t in transacoes if Decimal(t[2]) > 0]; await update.message.reply_text("Nenhuma saÃ­da.", reply_markup=teclado_flutuante(user_id)) if not filtradas else await update.message.reply_text("ğŸ’¸ SaÃ­das:\n" + "\n".join([f"â¬…ï¸ R$ {formatar_valor(t[2])} ({t[3]}) - {t[5] or 'Dinheiro'} - {t[6]}" for t in filtradas]), reply_markup=teclado_flutuante(user_id)); return

    if msg == "ğŸ—“ï¸ Filtrar por PerÃ­odo": 
        context.user_data['aguardando_filtro'] = True; await update.message.reply_text("Selecione o perÃ­odo:", reply_markup=teclado_filtros_periodo()); return

    # --- LÃ³gica de Filtro por Categoria (Atualizada com novo nome) ---
    if msg == "ğŸ·ï¸ Filtrar por Categoria":
        context.user_data['aguardando_filtro_categoria'] = True 
        trans_gastos = db.get_todas(user_id=user_id, tipo="gasto")
        trans_entradas = db.get_todas(user_id=user_id, tipo="entrada")
        cats_gasto = {t[3] for t in trans_gastos if Decimal(t[2]) > 0}
        cats_entrada = {t[3] for t in trans_entradas if Decimal(t[2]) > 0}
        categorias_unicas = sorted(list(cats_gasto.union(cats_entrada)))
        if not categorias_unicas:
            await update.message.reply_text("Nenhuma categoria registrada ainda.", reply_markup=teclado_flutuante(user_id))
            del context.user_data['aguardando_filtro_categoria']; return
        
        teclado_categorias = []; linha_atual = []
        for cat in categorias_unicas:
            linha_atual.append(cat)
            if len(linha_atual) == 2: teclado_categorias.append(linha_atual); linha_atual = []
        if linha_atual: teclado_categorias.append(linha_atual)
        teclado_categorias.append(["Cancelar"]) 
        await update.message.reply_text("Selecione uma categoria para filtrar:", reply_markup=ReplyKeyboardMarkup(teclado_categorias, resize_keyboard=True, one_time_keyboard=True)); return
    # --- Fim do Bloco Modificado ---

    if msg == "ğŸ’³ Gastos por CartÃ£o": 
        texto = gastos_por_cartao(user_id); await update.message.reply_text(texto, reply_markup=teclado_flutuante(user_id)); return
    if msg == "âš–ï¸ Saldo Geral":
        entradas = db.get_soma(user_id, "entrada"); gastos = db.get_soma(user_id, "gasto"); saldo = entradas - gastos
        status = "ğŸŸ¢ğŸ˜€ SaudÃ¡vel";
        if saldo < 0: status = "ğŸ”´ğŸ˜Ÿ Negativo"
        elif entradas > 0 and (gastos / entradas) > Decimal("0.7"): status = "ğŸŸ ğŸ¤” Gastos altos!"
        await update.message.reply_text((f"ğŸ§¾ Saldo Geral\nğŸ’° Entradas: R$ {formatar_valor(entradas)}\nğŸ’¸ Gastos: R$ {formatar_valor(gastos)}\nğŸ“Œ Saldo: R$ {formatar_valor(saldo)}\n\nStatus: {status}"), reply_markup=teclado_flutuante(user_id)); return

    if msg == "ğŸ“„ Gerar PDF": 
        filename = gerar_pdf(user_id); await update.message.reply_document(open(filename, "rb"), reply_markup=teclado_flutuante(user_id)); os.remove(filename); return
    if msg == "ğŸ“ˆ Gerar XLSX": 
        filename = gerar_xlsx(user_id); await update.message.reply_document(open(filename, "rb"), reply_markup=teclado_flutuante(user_id)); os.remove(filename); return

    # ==========================================================
    # --- MODIFICAÃ‡ÃƒO 3: Adicionar resposta "Quero um robÃ´" ---
    # ==========================================================
    if msg == "ğŸ¤– Quero um robÃ´":
        # Assumindo que seu username do Telegram Ã© 'maiconjbf'
        # Se for outro, troque o link abaixo.
        await update.message.reply_text(
            "Ã“tima ideia! Eu tambÃ©m posso criar um robÃ´ personalizado para vocÃª ou sua empresa.\n\n"
            "Me chame no Telegram para discutir seu projeto: ğŸ‘‰ https://t.me/maicon_junio",
            reply_markup=teclado_flutuante(user_id) # MantÃ©m o teclado principal
        )
        return
    # --- Fim do Novo Bloco ---

    # --- LÃ³gica Admin (Listar/Selecionar) ---
    if msg == "ğŸ§‘â€ğŸ’¼ Ver UsuÃ¡rios" and user_id == ADMIN_USER_ID: 
        usuarios = db.listar_usuarios(); await update.message.reply_text("Nenhum usuÃ¡rio.", reply_markup=teclado_flutuante(user_id)) if not usuarios else await update.message.reply_text("Gerenciar usuÃ¡rio:", reply_markup=ReplyKeyboardMarkup([[f"{u[0]} - {u[1]}"] for u in usuarios] + [["â¬…ï¸ Voltar"]], resize_keyboard=True, one_time_keyboard=True)); return
    if user_id == ADMIN_USER_ID and " - " in msg and msg.split(" - ")[0].isdigit(): 
        selecionado_id = int(msg.split(" - ")[0]); selecionado_nome = msg.split(" - ")[1]; context.user_data["admin_selecionado"] = (selecionado_id, selecionado_nome); await update.message.reply_text(f"Gerenciando: {selecionado_nome}.", reply_markup=teclado_admin_usuario_selecionado()); return

    # --- InterpretaÃ§Ã£o de Mensagem (Adicionar transaÃ§Ã£o) ---
    resultado = interpretar_mensagem(msg)
    if resultado["acao"] == "add":
        db.add_transacao(user_id, resultado["tipo"], resultado["valor_num"], resultado["valor_txt"], resultado["categoria"], resultado["metodo"], resultado["cartao"], user_name)
        msg_resp = f"âœ… {resultado['tipo'].capitalize()} R$ {formatar_valor(resultado['valor_num'])} (Cat: {resultado['categoria']})"
        if resultado['cartao']: msg_resp += f"\nğŸ’³ CartÃ£o: {resultado['cartao']}"
        alerta = verificar_alerta(user_id)
        if alerta: msg_resp += f"\n\n{alerta}"
        await update.message.reply_text(msg_resp, reply_markup=teclado_flutuante(user_id))
    else:
        await update.message.reply_text("âŒ NÃ£o entendi. Digite valor + descriÃ§Ã£o (ex: '50 lanche').", reply_markup=teclado_flutuante(user_id))

# =======================
# InicializaÃ§Ã£o do Bot (VERSÃƒO FINAL E CORRIGIDA - Sem alteraÃ§Ãµes aqui)
# =======================
TOKEN = os.environ.get('BOT_TOKEN') # LÃª o token dos "Secrets"
app = None # VariÃ¡vel global para o app do Telegram
if not TOKEN: print("ERRO CRÃTICO: Token nÃ£o encontrado.")
else:
    app = Application.builder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, responder))
    print("ğŸ¤– Bot configurado.")

# --- CÃ“DIGO DO SERVIDOR FLASK (Sem alteraÃ§Ãµes aqui) ---
app_flask = Flask('')
@app_flask.route('/')
def home(): return "Estou vivo!"

def run_telegram_bot():
    print("ğŸ¤– Bot do Telegram rodando em background.")
    try: app.run_polling(stop_signals=None)
    except Exception as e: print(f"!!! ERRO FATAL NO POLLING: {e} !!!")

def run_flask_and_bot():
    if not app: return
    print("Iniciando Bot e Servidor...")
    thread_bot = Thread(target=run_telegram_bot, daemon=True)
    thread_bot.start()
    print("\n--- INICIANDO FLASK ---")
    try:
        from waitress import serve
        print("--- Waitress na porta 8080 ---")
        serve(app_flask, host='0.0.0.0', port=8080)
    except ImportError:
        print("--- Fallback Flask na porta 8080 ---")
        app_flask.run(host='0.0.0.0', port=8080)
    except Exception as e: print(f"!!! ERRO FLASK: {e} !!!")

if __name__ == "__main__":
    run_flask_and_bot()